"""
Read data-raw/oseberg/<latest>/completions_wab.xlsx, normalize, and write
data/completions.json. Updates only completions-related counters in
data/meta.json.

See docs/2026-05-22-data-model-spec.md §7.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import CANONICAL_COUNTIES, NormalizationError, normalize_county  # noqa: E402
from ingest_oseberg_wells import (  # noqa: E402
    find_latest_oseberg_folder,
    parse_section_from_legal,
    to_float_or_none,
    to_int_or_none,
    to_str_or_none,
    _date_or_none,
)
from openpyxl import load_workbook  # noqa: E402

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
COMPLETIONS_OUT = DATA_DIR / "completions.json"
META_OUT = DATA_DIR / "meta.json"


def load_owned_index(tracts_path: Path) -> Dict[str, List[str]]:
    payload = json.load(tracts_path.open("r", encoding="utf-8"))
    idx: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        idx.setdefault(t["str"], []).append(t["tract_id"])
    for k in idx:
        idx[k].sort()
    return idx


def parse_completions(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], dict, List[dict]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    required = [
        "County", "API Number", "Operator", "Well Name", "Legal",
        "Completion Date", "Recompletion Date", "Recomplete", "Type Of Document",
        "Formations", "IP Oil (BOPD)", "IP Gas (MCFPD)", "IP Water (BWPD)",
        "Bottom Hole Total Length", "Effective Date", "State Url",
        "Latitude", "Longitude",
    ]
    missing = [c for c in required if c not in H]
    if missing:
        raise RuntimeError(f"completions_wab.xlsx missing expected columns: {missing}")

    rows_out: List[dict] = []
    ingestion_errors: List[dict] = []
    dropped_out_of_scope = 0
    affecting_owned = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county_raw = row[H["County"]]
        if county_raw is None:
            continue
        try:
            county = normalize_county(county_raw)
        except NormalizationError:
            dropped_out_of_scope += 1
            continue
        if county not in CANONICAL_COUNTIES:
            dropped_out_of_scope += 1
            continue

        api_number = to_str_or_none(row[H["API Number"]])
        completion_date = _date_or_none(row[H["Completion Date"]])
        recompletion_date = _date_or_none(row[H["Recompletion Date"]])
        effective_date = _date_or_none(row[H["Effective Date"]])
        recomplete_flag = to_str_or_none(row[H["Recomplete"]])
        is_recomp = recomplete_flag and recomplete_flag.upper() == "Y"

        date_for_id = completion_date or recompletion_date or effective_date or "undated"
        comp_id_key = api_number or f"{county.lower().replace(' ', '-')}-r{r_idx}"
        completion_id = f"comp-{comp_id_key}-{date_for_id}"

        legal_raw = row[H["Legal"]]
        section_norm = parse_section_from_legal(legal_raw)
        sections = [section_norm] if section_norm else []

        matched_tract_ids = sorted(set(owned_index.get(sections[0], [])) if sections else set())
        if matched_tract_ids:
            affecting_owned += 1

        record = {
            "completion_date": completion_date,
            "completion_id": completion_id,
            "completion_type": "Recompletion" if is_recomp else "Initial",
            "county": county,
            "effective_date": effective_date,
            "formation": to_str_or_none(row[H["Formations"]]),
            "ip_gas_mcfpd": to_float_or_none(row[H["IP Gas (MCFPD)"]]),
            "ip_oil_bopd": to_float_or_none(row[H["IP Oil (BOPD)"]]),
            "ip_water_bwpd": to_float_or_none(row[H["IP Water (BWPD)"]]),
            "lat": to_float_or_none(row[H["Latitude"]]),
            "lateral_length_ft": to_int_or_none(row[H["Bottom Hole Total Length"]]),
            "lon": to_float_or_none(row[H["Longitude"]]),
            "matched_tract_ids": matched_tract_ids,
            "operator": to_str_or_none(row[H["Operator"]]),
            "oseberg_url": to_str_or_none(row[H["State Url"]]),
            "recompletion_date": recompletion_date,
            "sections": sections,
            "type_of_document": to_str_or_none(row[H["Type Of Document"]]),
            "well_id": api_number,
            "well_name": to_str_or_none(row[H["Well Name"]]),
        }
        rows_out.append(record)

    counters = {
        "completions_dropped_out_of_scope": dropped_out_of_scope,
        "completions_affecting_owned_tracts": affecting_owned,
    }
    return rows_out, counters, ingestion_errors


def merge_meta(meta_path: Path, *, source_folder: str, count: int, counters: dict, errors: List[dict]) -> dict:
    if meta_path.exists():
        meta = json.load(meta_path.open("r", encoding="utf-8"))
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}
    meta.setdefault("counts", {})["completions"] = count
    mr = meta.setdefault("matching_report", {})
    mr["completions_affecting_owned_tracts"] = counters["completions_affecting_owned_tracts"]
    mr["completions_dropped_out_of_scope"] = counters["completions_dropped_out_of_scope"]
    existing_errors = mr.get("ingestion_errors", [])
    mr["ingestion_errors"] = [e for e in existing_errors if e.get("section") != "completions"] + errors
    meta["generated_at"] = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not meta.get("oseberg_folder"):
        meta["oseberg_folder"] = source_folder
    return meta


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oseberg-folder", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    folder = OSEBERG_ROOT / args.oseberg_folder if args.oseberg_folder else find_latest_oseberg_folder(OSEBERG_ROOT)
    xlsx = folder / "completions_wab.xlsx"
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr)
        return 2

    owned_index = load_owned_index(TRACTS_IN)
    records, counters, errors = parse_completions(xlsx, owned_index)
    records.sort(
        key=lambda r: (r["completion_date"] or r["effective_date"] or "0000-00-00", r["completion_id"]),
        reverse=True,
    )

    payload = {
        "generated_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source": f"oseberg-{folder.name}",
        "completions": records,
    }

    if not args.dry_run:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        json.dump(payload, COMPLETIONS_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        COMPLETIONS_OUT.open("a", encoding="utf-8").write("\n")
        meta = merge_meta(META_OUT, source_folder=folder.name, count=len(records), counters=counters, errors=errors)
        json.dump(meta, META_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        META_OUT.open("a", encoding="utf-8").write("\n")

    print(f"Completions in scope:               {len(records)}")
    print(f"  affecting owned tracts:           {counters['completions_affecting_owned_tracts']}")
    print(f"Completions dropped (out of scope): {counters['completions_dropped_out_of_scope']}")
    print(f"Ingestion errors:                   {len(errors)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
