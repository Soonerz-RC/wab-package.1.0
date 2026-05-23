"""
Read data-raw/oseberg/<latest>/leasing_wab.xlsx and write two outputs per
spec §9.3:
  - data/leasing.json         : leases where affects_owned_tracts == true
  - data/leasing_market.json  : full WAB market set (loaded on demand)

See docs/2026-05-22-data-model-spec.md §9.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import (  # noqa: E402
    CANONICAL_COUNTIES,
    NormalizationError,
    normalize_county,
    normalize_upland_in_string,
)
from ingest_oseberg_wells import (  # noqa: E402
    find_latest_oseberg_folder,
    parse_section_from_legal,
    to_float_or_none,
    to_int_or_none,
    to_str_or_none,
    _date_or_none,
)
from openpyxl import load_workbook  # noqa: E402

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
LEASING_OWNED_OUT = DATA_DIR / "leasing.json"
LEASING_MARKET_OUT = DATA_DIR / "leasing_market.json"
META_OUT = DATA_DIR / "meta.json"


def load_owned_index(tracts_path: Path) -> Dict[str, List[str]]:
    payload = json.load(tracts_path.open("r", encoding="utf-8"))
    idx: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        idx.setdefault(t["str"], []).append(t["tract_id"])
    for k in idx:
        idx[k].sort()
    return idx


def _term_to_years(term, unit) -> Optional[float]:
    """Convert (Term, Term Unit) into years. Unit is typically 'Years' or 'Months'."""
    t = to_float_or_none(term)
    if t is None:
        return None
    u = to_str_or_none(unit) or ""
    u_lower = u.lower()
    if "year" in u_lower:
        return round(t, 2)
    if "month" in u_lower:
        return round(t / 12.0, 2)
    if "day" in u_lower:
        return round(t / 365.0, 2)
    # Unknown unit — return raw and let the consumer decide
    return round(t, 2)


def parse_leasing(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], List[dict], dict, List[dict]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    required = [
        "County", "Legal", "Classification", "Instrument Type", "Recorded Date",
        "Effective Date", "Term", "Term Unit", "Expiration Date",
        "Lessor/Grantor", "Lessee/Grantee", "Acreage", "Royalty",
        "Book", "Number", "source_url",
    ]
    missing = [c for c in required if c not in H]
    if missing:
        raise RuntimeError(f"leasing_wab.xlsx missing expected columns: {missing}")

    owned_only: List[dict] = []
    market: List[dict] = []
    ingestion_errors: List[dict] = []
    dropped_out_of_scope = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county_raw = row[H["County"]]
        if county_raw is None:
            dropped_out_of_scope += 1
            continue
        try:
            county = normalize_county(county_raw)
        except NormalizationError:
            dropped_out_of_scope += 1
            continue
        if county not in CANONICAL_COUNTIES:
            dropped_out_of_scope += 1
            continue

        legal_raw = to_str_or_none(row[H["Legal"]])
        section_norm = parse_section_from_legal(legal_raw)
        sections = [section_norm] if section_norm else []

        matched_tract_ids: List[str] = []
        if sections:
            for tid in owned_index.get(sections[0], []):
                matched_tract_ids.append(tid)
        matched_tract_ids.sort()
        affects = bool(matched_tract_ids)

        book = to_str_or_none(row[H["Book"]])
        number = to_str_or_none(row[H["Number"]])
        lease_id_parts = [s for s in (county.lower().replace(" ", "-"), book, number) if s]
        lease_id = "lease-" + "-".join(lease_id_parts) if lease_id_parts else f"lease-r{r_idx}"

        record = {
            "acres": to_float_or_none(row[H["Acreage"]]),
            "affects_owned_tracts": affects,
            "book_page": f"{book}/{number}" if book and number else (book or number),
            "classification": to_str_or_none(row[H["Classification"]]),
            "county": county,
            "instrument_date": _date_or_none(row[H["Effective Date"]]),
            "instrument_type": to_str_or_none(row[H["Instrument Type"]]),
            "legal_raw": legal_raw,
            "lease_id": lease_id,
            "lessee": normalize_upland_in_string(to_str_or_none(row[H["Lessee/Grantee"]])),
            "lessor": normalize_upland_in_string(to_str_or_none(row[H["Lessor/Grantor"]])),
            "matched_tract_ids": matched_tract_ids,
            "oseberg_url": to_str_or_none(row[H["source_url"]]),
            "recording_date": _date_or_none(row[H["Recorded Date"]]),
            "royalty": to_float_or_none(row[H["Royalty"]]),
            "sections": sections,
            "term_expiration_date": _date_or_none(row[H["Expiration Date"]]),
            "term_years": _term_to_years(row[H["Term"]], row[H["Term Unit"]]),
        }
        market.append(record)
        if affects:
            owned_only.append(record)

    counters = {
        "leases_affecting_owned_tracts": len(owned_only),
        "leasing_market_count": len(market),
        "leasing_dropped_out_of_scope": dropped_out_of_scope,
    }
    return owned_only, market, counters, ingestion_errors


def _write_json(path: Path, payload: dict) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, sort_keys=True, ensure_ascii=False)
        f.write("\n")


def merge_meta(meta_path: Path, *, source_folder: str, counters: dict, errors: List[dict]) -> dict:
    if meta_path.exists():
        meta = json.load(meta_path.open("r", encoding="utf-8"))
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}
    meta.setdefault("counts", {})["leases"] = counters["leases_affecting_owned_tracts"]
    mr = meta.setdefault("matching_report", {})
    mr["leases_affecting_owned_tracts"] = counters["leases_affecting_owned_tracts"]
    mr["leasing_market_count"] = counters["leasing_market_count"]
    mr["leasing_dropped_out_of_scope"] = counters["leasing_dropped_out_of_scope"]
    existing_errors = mr.get("ingestion_errors", [])
    mr["ingestion_errors"] = [e for e in existing_errors if e.get("section") != "leasing"] + errors
    meta["generated_at"] = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not meta.get("oseberg_folder"):
        meta["oseberg_folder"] = source_folder
    return meta


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oseberg-folder", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    folder = OSEBERG_ROOT / args.oseberg_folder if args.oseberg_folder else find_latest_oseberg_folder(OSEBERG_ROOT)
    xlsx = folder / "leasing_wab.xlsx"
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr)
        return 2

    owned_index = load_owned_index(TRACTS_IN)
    print(f"Parsing {xlsx.name} (large file — may take 30-60s)...", flush=True)
    owned_only, market, counters, errors = parse_leasing(xlsx, owned_index)

    def _sort_key(r):
        # Recording date desc, then lease_id for stability
        return ("0000-00-00" if not r["recording_date"] else r["recording_date"], r["lease_id"])

    owned_only.sort(key=_sort_key, reverse=True)
    market.sort(key=_sort_key, reverse=True)

    now_utc = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    owned_payload = {
        "generated_at": now_utc,
        "source": f"oseberg-{folder.name}",
        "scope": "leases affecting owned tracts (per spec §9.3)",
        "leases": owned_only,
    }
    market_payload = {
        "generated_at": now_utc,
        "source": f"oseberg-{folder.name}",
        "scope": "full WAB market lease set (per spec §9.3)",
        "leases": market,
    }

    if not args.dry_run:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        _write_json(LEASING_OWNED_OUT, owned_payload)
        _write_json(LEASING_MARKET_OUT, market_payload)
        meta = merge_meta(META_OUT, source_folder=folder.name, counters=counters, errors=errors)
        _write_json(META_OUT, meta)

    print(f"Leases (owned-affecting):  {len(owned_only)}")
    print(f"Leases (full WAB market):  {len(market)}")
    print(f"Dropped out-of-scope:      {counters['leasing_dropped_out_of_scope']}")
    print(f"leasing.json:              {LEASING_OWNED_OUT.stat().st_size / 1024:.1f} KB" if LEASING_OWNED_OUT.exists() else "leasing.json: (not written)")
    print(f"leasing_market.json:       {LEASING_MARKET_OUT.stat().st_size / 1_000_000:.2f} MB" if LEASING_MARKET_OUT.exists() else "leasing_market.json: (not written)")
    print(f"Ingestion errors:          {len(errors)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
