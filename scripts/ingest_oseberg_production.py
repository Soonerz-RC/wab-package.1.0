"""
Read data-raw/oseberg/<latest>/production_wab.xlsx, normalize, and write
data/production.json.

The Oseberg production export is lease-level lifetime summary data (one
record per producing lease unit), not monthly well-level production.
See docs/2026-05-22-data-model-spec.md §8 for the schema.

The single most important file for demonstrating HBP (Held By Production)
status — every lease with a recent last_prod_date + meaningful cumulative
volumes is currently holding its underlying mineral position.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import CANONICAL_COUNTIES, NormalizationError, normalize_county  # noqa: E402
from ingest_oseberg_wells import (  # noqa: E402
    find_latest_oseberg_folder,
    parse_section_from_legal,
    to_float_or_none,
    to_int_or_none,
    to_str_or_none,
    _date_or_none,
)
from openpyxl import load_workbook  # noqa: E402

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
PRODUCTION_OUT = DATA_DIR / "production.json"
META_OUT = DATA_DIR / "meta.json"

# Threshold for "is_active": within this many months of generated_at
ACTIVE_THRESHOLD_MONTHS = 12


def load_owned_index(tracts_path: Path) -> Dict[str, List[str]]:
    payload = json.load(tracts_path.open("r", encoding="utf-8"))
    idx: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        idx.setdefault(t["str"], []).append(t["tract_id"])
    for k in idx:
        idx[k].sort()
    return idx


def split_api_cell(value) -> List[str]:
    """Oseberg packs multiple APIs into one cell, semicolon-separated."""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    return [p.strip() for p in s.split(";") if p.strip()]


def _is_active(last_prod_iso: Optional[str], today: dt.date) -> bool:
    if not last_prod_iso:
        return False
    try:
        d = dt.date.fromisoformat(last_prod_iso)
    except ValueError:
        return False
    months_ago = (today.year - d.year) * 12 + (today.month - d.month)
    return months_ago <= ACTIVE_THRESHOLD_MONTHS


def parse_production(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], dict, List[dict]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    required = [
        "County", "Legal", "Operator", "Lease Name", "API Number", "Lease Number",
        "Reservoir Name", "Field Name", "Active Date", "First Prod Date",
        "Last Prod Date", "Latest Completion Date", "Number of Months Producing",
        "Number of Completions", "Cumulative Oil (BBL)", "Cumulative Gas (MCF)",
        "Last Month Production Oil (BOPM)", "Last Month Production Gas (MCFPM)",
        "Month Over Month Oil", "Month Over Month Gas",
        "Year Over Year Oil", "Year Over Year Gas",
        "IP30 Oil (BOPD)", "IP30 Gas (MCFPD)",
        "IP60 Oil (BOPD)", "IP60 Gas (MCFPD)",
        "IP90 Oil (BOPD)", "IP90 Gas (MCFPD)",
        "First Completion IP Oil", "First Completion IP Gas", "First Completion IP Water",
        "Latest Completion IP Oil", "Latest Completion IP Gas", "Latest Completion IP Water",
        "Best 30 Oil (BOPM)", "Best 30 Gas (MCFPM)",
        "Avg Last 12 Month Oil (BOPM)", "Avg Last 12 Month Gas (MCFPM)",
        "Sum Last 12 Month Oil (BOPM)", "Sum Last 12 Month Gas (MCFPM)",
        "Decline Rate Oil", "Decline Rate Gas",
        "Sum of Lateral Length", "Gross Acres", "lease_unit_id",
    ]
    missing = [c for c in required if c not in H]
    if missing:
        raise RuntimeError(f"production_wab.xlsx missing expected columns: {missing}")

    today = dt.date.today()
    records: List[dict] = []
    ingestion_errors: List[dict] = []
    dropped_out_of_scope = 0
    affecting_owned = 0

    def slug(s):
        if not s:
            return "unknown"
        return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-") or "unknown"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county_raw = row[H["County"]]
        if county_raw is None:
            continue
        try:
            county = normalize_county(county_raw)
        except NormalizationError:
            dropped_out_of_scope += 1
            continue
        if county not in CANONICAL_COUNTIES:
            dropped_out_of_scope += 1
            continue

        lease_number = to_str_or_none(row[H["Lease Number"]])
        lease_name = to_str_or_none(row[H["Lease Name"]])
        operator = to_str_or_none(row[H["Operator"]])
        legal_raw = to_str_or_none(row[H["Legal"]])
        section_norm = parse_section_from_legal(legal_raw)
        sections = [section_norm] if section_norm else []

        matched_tract_ids = sorted(set(owned_index.get(sections[0], [])) if sections else set())
        if matched_tract_ids:
            affecting_owned += 1

        api_numbers = split_api_cell(row[H["API Number"]])
        last_prod_date = _date_or_none(row[H["Last Prod Date"]])

        # Build production_id
        if lease_number:
            production_id = f"lease-{lease_number}"
        else:
            production_id = f"lease-{slug(county)}-{section_norm or 'unknown'}-{slug(operator)}-r{r_idx}"

        record = {
            "active_date": _date_or_none(row[H["Active Date"]]),
            "affects_owned_tracts": bool(matched_tract_ids),
            "api_numbers": api_numbers,
            "avg_last_12_month_gas_mcfpm": to_float_or_none(row[H["Avg Last 12 Month Gas (MCFPM)"]]),
            "avg_last_12_month_oil_bopm": to_float_or_none(row[H["Avg Last 12 Month Oil (BOPM)"]]),
            "best_30_gas_mcfpm": to_float_or_none(row[H["Best 30 Gas (MCFPM)"]]),
            "best_30_oil_bopm": to_float_or_none(row[H["Best 30 Oil (BOPM)"]]),
            "county": county,
            "cumulative_gas_mcf": to_float_or_none(row[H["Cumulative Gas (MCF)"]]),
            "cumulative_oil_bbl": to_float_or_none(row[H["Cumulative Oil (BBL)"]]),
            "decline_rate_gas": to_float_or_none(row[H["Decline Rate Gas"]]),
            "decline_rate_oil": to_float_or_none(row[H["Decline Rate Oil"]]),
            "field_name": to_str_or_none(row[H["Field Name"]]),
            "first_completion_ip_gas": to_float_or_none(row[H["First Completion IP Gas"]]),
            "first_completion_ip_oil": to_float_or_none(row[H["First Completion IP Oil"]]),
            "first_completion_ip_water": to_float_or_none(row[H["First Completion IP Water"]]),
            "first_prod_date": _date_or_none(row[H["First Prod Date"]]),
            "gross_acres": to_float_or_none(row[H["Gross Acres"]]),
            "ip30_gas_mcfpd": to_float_or_none(row[H["IP30 Gas (MCFPD)"]]),
            "ip30_oil_bopd": to_float_or_none(row[H["IP30 Oil (BOPD)"]]),
            "ip60_gas_mcfpd": to_float_or_none(row[H["IP60 Gas (MCFPD)"]]),
            "ip60_oil_bopd": to_float_or_none(row[H["IP60 Oil (BOPD)"]]),
            "ip90_gas_mcfpd": to_float_or_none(row[H["IP90 Gas (MCFPD)"]]),
            "ip90_oil_bopd": to_float_or_none(row[H["IP90 Oil (BOPD)"]]),
            "is_active": _is_active(last_prod_date, today),
            "last_month_gas_mcfpm": to_float_or_none(row[H["Last Month Production Gas (MCFPM)"]]),
            "last_month_oil_bopm": to_float_or_none(row[H["Last Month Production Oil (BOPM)"]]),
            "last_prod_date": last_prod_date,
            "latest_completion_date": _date_or_none(row[H["Latest Completion Date"]]),
            "latest_completion_ip_gas": to_float_or_none(row[H["Latest Completion IP Gas"]]),
            "latest_completion_ip_oil": to_float_or_none(row[H["Latest Completion IP Oil"]]),
            "latest_completion_ip_water": to_float_or_none(row[H["Latest Completion IP Water"]]),
            "lateral_length_sum_ft": to_int_or_none(row[H["Sum of Lateral Length"]]),
            "legal_raw": legal_raw,
            "lease_name": lease_name,
            "lease_number": lease_number,
            "matched_tract_ids": matched_tract_ids,
            "month_over_month_gas": to_float_or_none(row[H["Month Over Month Gas"]]),
            "month_over_month_oil": to_float_or_none(row[H["Month Over Month Oil"]]),
            "number_of_completions": to_int_or_none(row[H["Number of Completions"]]),
            "number_of_months_producing": to_int_or_none(row[H["Number of Months Producing"]]),
            "operator": operator,
            "production_id": production_id,
            "reservoir_name": to_str_or_none(row[H["Reservoir Name"]]),
            "sections": sections,
            "sum_last_12_month_gas_mcfpm": to_float_or_none(row[H["Sum Last 12 Month Gas (MCFPM)"]]),
            "sum_last_12_month_oil_bopm": to_float_or_none(row[H["Sum Last 12 Month Oil (BOPM)"]]),
            "well_ids": api_numbers,
            "year_over_year_gas": to_float_or_none(row[H["Year Over Year Gas"]]),
            "year_over_year_oil": to_float_or_none(row[H["Year Over Year Oil"]]),
        }
        records.append(record)

    counters = {
        "production_dropped_out_of_scope": dropped_out_of_scope,
        "production_affecting_owned_tracts": affecting_owned,
    }
    return records, counters, ingestion_errors


def merge_meta(meta_path: Path, *, source_folder: str, count: int, counters: dict, errors: List[dict]) -> dict:
    if meta_path.exists():
        meta = json.load(meta_path.open("r", encoding="utf-8"))
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}
    meta.setdefault("counts", {})["production_records"] = count
    mr = meta.setdefault("matching_report", {})
    mr["production_affecting_owned_tracts"] = counters["production_affecting_owned_tracts"]
    mr["production_dropped_out_of_scope"] = counters["production_dropped_out_of_scope"]
    existing_errors = mr.get("ingestion_errors", [])
    mr["ingestion_errors"] = [e for e in existing_errors if e.get("section") != "production"] + errors
    meta["generated_at"] = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not meta.get("oseberg_folder"):
        meta["oseberg_folder"] = source_folder
    return meta


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oseberg-folder", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    folder = OSEBERG_ROOT / args.oseberg_folder if args.oseberg_folder else find_latest_oseberg_folder(OSEBERG_ROOT)
    xlsx = folder / "production_wab.xlsx"
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr)
        return 2

    owned_index = load_owned_index(TRACTS_IN)
    records, counters, errors = parse_production(xlsx, owned_index)
    # Sort: owned-affecting first, then by last_prod_date desc
    records.sort(key=lambda r: (
        not r["affects_owned_tracts"],
        -(int(r["last_prod_date"][:4]) * 100 + int(r["last_prod_date"][5:7])) if r["last_prod_date"] else 0,
        r["production_id"],
    ))

    payload = {
        "generated_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source": f"oseberg-{folder.name}",
        "production": records,
    }

    if not args.dry_run:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        json.dump(payload, PRODUCTION_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        PRODUCTION_OUT.open("a", encoding="utf-8").write("\n")
        meta = merge_meta(META_OUT, source_folder=folder.name, count=len(records), counters=counters, errors=errors)
        json.dump(meta, META_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        META_OUT.open("a", encoding="utf-8").write("\n")

    print(f"Production records in scope:        {len(records)}")
    print(f"  affecting owned tracts:           {counters['production_affecting_owned_tracts']}")
    print(f"Production dropped (out of scope):  {counters['production_dropped_out_of_scope']}")
    print(f"Active leases (last prod ≤ 12 mo): {sum(1 for r in records if r['is_active'])}")
    print(f"Ingestion errors:                   {len(errors)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
