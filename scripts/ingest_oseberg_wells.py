"""
Read data-raw/oseberg/<latest>/wells_wab.xlsx, normalize, compute
sections[] from {Surface Hole Legal, Bottom Hole Legal} endpoints,
match against owned tracts, and write data/wells.json. Updates the
wells-related counters in data/meta.json without disturbing other
counters (additive merge).

See docs/2026-05-22-data-model-spec.md §5 (and §5.4 / §5.5 for the
endpoint-based sections logic and column mapping).
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Sibling import
SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import (  # noqa: E402
    CANONICAL_COUNTIES,
    NormalizationError,
    normalize_county,
    normalize_str,
    to_iso_date_or_hbp,
)

from openpyxl import load_workbook  # noqa: E402


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
WELLS_OUT = DATA_DIR / "wells.json"
META_OUT = DATA_DIR / "meta.json"

# Long-lateral flagging threshold (one mile in feet). Wells with lateral length
# at or above this can cross more than 2 sections per spec §5.4.
LONG_LATERAL_THRESHOLD_FT = 5280


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def find_latest_oseberg_folder(root: Path) -> Path:
    """Return the most recent dated subfolder under data-raw/oseberg/.

    Subfolders are expected to be named ``YYYY-MM-DD``. The function ignores
    anything that doesn't parse as an ISO date.
    """
    if not root.exists():
        raise FileNotFoundError(f"Oseberg root not found: {root}")
    candidates: List[Tuple[dt.date, Path]] = []
    for p in root.iterdir():
        if not p.is_dir():
            continue
        try:
            d = dt.date.fromisoformat(p.name)
        except ValueError:
            continue
        candidates.append((d, p))
    if not candidates:
        raise FileNotFoundError(
            f"No dated subfolders in {root}. Expected at least one YYYY-MM-DD/"
        )
    candidates.sort(reverse=True)
    return candidates[0][1]


def strip_meridian_suffix(legal: Optional[str]) -> Optional[str]:
    """Remove the trailing ``-IM`` (or similar 2-letter meridian) from a legal string.

    Examples:
        "31-12N-10W-IM"  -> "31-12N-10W"
        "18-17N-24W-IM"  -> "18-17N-24W"
        "31-12N-10W"     -> "31-12N-10W"  (no suffix to strip)
        None             -> None
    """
    if legal is None:
        return None
    s = str(legal).strip()
    if not s:
        return None
    # Match trailing "-XX" where XX is 1-3 alphabetic chars (meridian code)
    return re.sub(r"-[A-Z]{1,3}$", "", s.upper())


def parse_section_from_legal(legal: Optional[str]) -> Optional[str]:
    """Parse a legal string into a canonical STR, returning None on failure.

    Does NOT raise — callers expect None for unparseable values so the well
    can still be included with whatever endpoint(s) did parse.
    """
    stripped = strip_meridian_suffix(legal)
    if stripped is None:
        return None
    try:
        return normalize_str(stripped)
    except NormalizationError:
        return None


def to_float_or_none(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in {"N/A", "NA", "-", "—"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def to_int_or_none(v) -> Optional[int]:
    f = to_float_or_none(v)
    return int(round(f)) if f is not None else None


def to_str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Wells ingest
# ---------------------------------------------------------------------------


def load_owned_str_set(tracts_path: Path) -> Dict[str, List[str]]:
    """Return {canonical_str: [tract_id, ...]} from tracts.json."""
    if not tracts_path.exists():
        raise FileNotFoundError(
            f"tracts.json not found at {tracts_path}. Run ingest_inventory.py first."
        )
    with tracts_path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    index: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        index.setdefault(t["str"], []).append(t["tract_id"])
    # Sort tract IDs within each section for deterministic matched_tract_ids
    for k in index:
        index[k].sort()
    return index


def parse_wells(
    wells_xlsx: Path,
    owned_index: Dict[str, List[str]],
) -> Tuple[List[dict], dict, List[dict]]:
    """Parse wells_wab.xlsx and return (wells_in_scope, counters, ingestion_errors).

    counters has:
      wells_dropped_out_of_scope, wells_with_owned_tract, wells_without_owned_tract,
      wells_long_laterals_flagged (list of dicts)
    """
    wb = load_workbook(wells_xlsx, data_only=True, read_only=True)
    if "Sheet 1" not in wb.sheetnames:
        # Fall back to first sheet
        sheet_name = wb.sheetnames[0]
    else:
        sheet_name = "Sheet 1"
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    H = {h: i for i, h in enumerate(header_row) if h is not None}

    required = [
        "API Number", "API Number 12 Digit", "Operator", "Original Operator",
        "Well Name", "County", "Well Status", "Well Type", "Wellbore Profile",
        "Spud Date", "Latest Completion Date",
        "Surface Latitude", "Surface Longitude",
        "Surface Hole Legal", "Bottom Hole Legal",
        "Lateral Length (Ft)", "State URL",
    ]
    missing = [c for c in required if c not in H]
    if missing:
        raise RuntimeError(f"wells_wab.xlsx missing expected columns: {missing}")

    wells: List[dict] = []
    ingestion_errors: List[dict] = []
    long_laterals: List[dict] = []
    dropped_out_of_scope = 0
    with_owned = 0
    without_owned = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county_raw = row[H["County"]]
        # Pre-filter on county before doing any heavier work
        if county_raw is None:
            ingestion_errors.append(
                {"section": "wells", "row": r_idx, "reason": "missing county", "raw_api": row[H["API Number"]]}
            )
            continue
        if county_raw not in CANONICAL_COUNTIES:
            # Try lenient title-case match
            try:
                _ = normalize_county(county_raw)
            except NormalizationError:
                dropped_out_of_scope += 1
                continue
        try:
            county = normalize_county(county_raw)
        except NormalizationError as exc:
            ingestion_errors.append(
                {"section": "wells", "row": r_idx, "reason": str(exc), "raw_county": county_raw}
            )
            continue

        api_10 = to_str_or_none(row[H["API Number"]])
        if api_10 is None:
            ingestion_errors.append(
                {"section": "wells", "row": r_idx, "reason": "missing API Number", "raw_county": county_raw}
            )
            continue

        well_id = api_10  # 10-digit API as the stable internal ID
        api_12 = to_str_or_none(row[H["API Number 12 Digit"]])

        shl_raw = row[H["Surface Hole Legal"]]
        bhl_raw = row[H["Bottom Hole Legal"]]
        shl_norm = parse_section_from_legal(shl_raw)
        bhl_norm = parse_section_from_legal(bhl_raw)

        if shl_raw and shl_norm is None:
            ingestion_errors.append(
                {
                    "section": "wells",
                    "row": r_idx,
                    "well_id": well_id,
                    "reason": f"unparseable Surface Hole Legal: {shl_raw!r}",
                }
            )
        if bhl_raw and bhl_norm is None:
            ingestion_errors.append(
                {
                    "section": "wells",
                    "row": r_idx,
                    "well_id": well_id,
                    "reason": f"unparseable Bottom Hole Legal: {bhl_raw!r}",
                }
            )

        sections = sorted({s for s in (shl_norm, bhl_norm) if s})

        # matched_tract_ids: union of all tract IDs in any of the sections[]
        matched_tract_ids: List[str] = []
        seen_ids: set = set()
        for sec in sections:
            for tid in owned_index.get(sec, []):
                if tid not in seen_ids:
                    matched_tract_ids.append(tid)
                    seen_ids.add(tid)
        matched_tract_ids.sort()

        if matched_tract_ids:
            with_owned += 1
        else:
            without_owned += 1

        # Operator history: current + original; dedupe if identical
        operator = to_str_or_none(row[H["Operator"]])
        original_operator = to_str_or_none(row[H["Original Operator"]])
        op_history: List[dict] = []
        if original_operator:
            op_history.append({"effective_date": None, "operator": original_operator})
        if operator and operator != original_operator:
            op_history.append({"effective_date": None, "operator": operator})
        elif operator and not original_operator:
            op_history.append({"effective_date": None, "operator": operator})

        spud_date = _date_or_none(row[H["Spud Date"]])
        completion_date = _date_or_none(row[H["Latest Completion Date"]])
        lateral_length = to_int_or_none(row[H["Lateral Length (Ft)"]])

        # Long-lateral flag
        if (
            lateral_length is not None
            and lateral_length >= LONG_LATERAL_THRESHOLD_FT
            and shl_norm is not None
            and bhl_norm is not None
            and shl_norm != bhl_norm
        ):
            long_laterals.append(
                {
                    "well_id": well_id,
                    "lateral_length_ft": lateral_length,
                    "surface_section": shl_norm,
                    "bottom_hole_section": bhl_norm,
                    "operator": operator,
                }
            )

        well_obj = {
            "api_number": api_12,
            "completion_date": completion_date,
            "county": county,
            "lat": to_float_or_none(row[H["Surface Latitude"]]),
            "lateral_length_ft": lateral_length,
            "lon": to_float_or_none(row[H["Surface Longitude"]]),
            "matched_tract_ids": matched_tract_ids,
            "operator": operator,
            "operator_history": op_history,
            "oseberg_url": to_str_or_none(row[H["State URL"]]),
            "sections": sections,
            "spud_date": spud_date,
            "well_id": well_id,
            "well_name": to_str_or_none(row[H["Well Name"]]),
            "well_status": to_str_or_none(row[H["Well Status"]]),
            "well_type": to_str_or_none(row[H["Well Type"]]),
            "wellbore_profile": to_str_or_none(row[H["Wellbore Profile"]]),
        }
        wells.append(well_obj)

    counters = {
        "wells_dropped_out_of_scope": dropped_out_of_scope,
        "wells_with_owned_tract": with_owned,
        "wells_without_owned_tract": without_owned,
        "wells_long_laterals_flagged": long_laterals,
    }
    return wells, counters, ingestion_errors


def _date_or_none(v) -> Optional[str]:
    """Best-effort date conversion. Returns None if the value can't be parsed."""
    if v is None:
        return None
    try:
        return to_iso_date_or_hbp(v)
    except NormalizationError:
        return None


# ---------------------------------------------------------------------------
# Meta merge — additive
# ---------------------------------------------------------------------------


def merge_meta(
    meta_path: Path,
    *,
    oseberg_folder: str,
    wells_count: int,
    counters: dict,
    ingestion_errors: List[dict],
) -> dict:
    """Read existing meta.json, merge in wells-related fields without clobbering others."""
    if meta_path.exists():
        with meta_path.open("r", encoding="utf-8") as f:
            meta = json.load(f)
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}

    meta["oseberg_folder"] = oseberg_folder

    counts = meta.setdefault("counts", {})
    counts["wells"] = wells_count

    mr = meta.setdefault("matching_report", {})
    mr["wells_with_owned_tract"] = counters["wells_with_owned_tract"]
    mr["wells_without_owned_tract"] = counters["wells_without_owned_tract"]
    mr["wells_dropped_out_of_scope"] = counters["wells_dropped_out_of_scope"]
    mr["wells_long_laterals_flagged"] = counters["wells_long_laterals_flagged"]

    # Append wells ingestion errors to whatever is already there (keep both sections' errors)
    existing_errors = mr.get("ingestion_errors", [])
    # Replace only wells-section entries with the fresh run's results
    others = [e for e in existing_errors if e.get("section") != "wells"]
    mr["ingestion_errors"] = others + ingestion_errors

    # Touch generated_at
    meta["generated_at"] = (
        dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )
    return meta


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report counts; do not write wells.json / meta.json.",
    )
    parser.add_argument(
        "--oseberg-folder",
        type=str,
        default=None,
        help="Override the auto-detected dated Oseberg folder (e.g. 2026-05-22).",
    )
    args = parser.parse_args(argv)

    if args.oseberg_folder:
        oseberg_folder = OSEBERG_ROOT / args.oseberg_folder
        if not oseberg_folder.exists():
            print(f"ERROR: {oseberg_folder} not found", file=sys.stderr)
            return 2
    else:
        oseberg_folder = find_latest_oseberg_folder(OSEBERG_ROOT)

    wells_xlsx = oseberg_folder / "wells_wab.xlsx"
    if not wells_xlsx.exists():
        print(f"ERROR: {wells_xlsx} not found", file=sys.stderr)
        return 3

    owned_index = load_owned_str_set(TRACTS_IN)
    wells, counters, ingestion_errors = parse_wells(wells_xlsx, owned_index)

    # Sort wells deterministically by well_id (numeric API)
    wells.sort(key=lambda w: w["well_id"])

    now_utc = (
        dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )
    wells_payload = {
        "generated_at": now_utc,
        "source": f"oseberg-{oseberg_folder.name}",
        "wells": wells,
    }

    if args.dry_run:
        print("DRY RUN — no files written.")
    else:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with WELLS_OUT.open("w", encoding="utf-8") as f:
            json.dump(wells_payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")
        meta = merge_meta(
            META_OUT,
            oseberg_folder=oseberg_folder.name,
            wells_count=len(wells),
            counters=counters,
            ingestion_errors=ingestion_errors,
        )
        with META_OUT.open("w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")

    print("=" * 64)
    print(f"Source: {wells_xlsx.relative_to(PROJECT_ROOT)}")
    print(f"Wells in scope:          {len(wells)}")
    print(f"  with owned tract:      {counters['wells_with_owned_tract']}")
    print(f"  without owned tract:   {counters['wells_without_owned_tract']}")
    print(f"Wells dropped out-of-scope: {counters['wells_dropped_out_of_scope']}")
    print(f"Long laterals flagged (>= {LONG_LATERAL_THRESHOLD_FT} ft, crossing 2+ sections): {len(counters['wells_long_laterals_flagged'])}")
    print(f"Ingestion errors:        {len(ingestion_errors)}")
    if ingestion_errors[:5]:
        print("  First 5 errors:")
        for e in ingestion_errors[:5]:
            print(f"    row {e.get('row')}: {e.get('reason')}")
    print("=" * 64)

    return 0


if __name__ == "__main__":
    sys.exit(main())
