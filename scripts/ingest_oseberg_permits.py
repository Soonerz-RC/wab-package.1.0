"""
Read data-raw/oseberg/<latest>/permit_wab.xlsx, normalize, and write
data/permits.json. Updates only permits-related counters in data/meta.json.

See docs/2026-05-22-data-model-spec.md §6.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import CANONICAL_COUNTIES, NormalizationError, normalize_county  # noqa: E402
from ingest_oseberg_wells import (  # noqa: E402
    find_latest_oseberg_folder,
    parse_section_from_legal,
    to_float_or_none,
    to_str_or_none,
    _date_or_none,
)
from openpyxl import load_workbook  # noqa: E402

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
PERMITS_OUT = DATA_DIR / "permits.json"
META_OUT = DATA_DIR / "meta.json"


def load_owned_index(tracts_path: Path) -> Dict[str, List[str]]:
    if not tracts_path.exists():
        raise FileNotFoundError(f"{tracts_path} not found. Run ingest_inventory.py first.")
    payload = json.load(tracts_path.open("r", encoding="utf-8"))
    idx: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        idx.setdefault(t["str"], []).append(t["tract_id"])
    for k in idx:
        idx[k].sort()
    return idx


def parse_permits(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], dict, List[dict]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    required = [
        "County", "API Number", "Operator", "Well Name", "Wellbore Profile",
        "Legal", "Bottom Hole Legal", "Filed Date", "Approval Date", "Effective Date",
        "Drilling Permit Num", "Type Of Document", "Purpose Of Filing", "Approval Status",
        "Source", "Latitude", "Longitude",
    ]
    missing = [c for c in required if c not in H]
    if missing:
        raise RuntimeError(f"permit_wab.xlsx missing expected columns: {missing}")

    permits: List[dict] = []
    ingestion_errors: List[dict] = []
    dropped_out_of_scope = 0
    affecting_owned = 0
    derived_id_counter: Dict[str, int] = {}

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county_raw = row[H["County"]]
        if county_raw is None:
            ingestion_errors.append(
                {"section": "permits", "row": r_idx, "reason": "missing county"}
            )
            continue
        try:
            county = normalize_county(county_raw)
        except NormalizationError:
            dropped_out_of_scope += 1
            continue
        if county not in CANONICAL_COUNTIES:
            dropped_out_of_scope += 1
            continue

        permit_number = to_str_or_none(row[H["Drilling Permit Num"]])
        api_number = to_str_or_none(row[H["API Number"]])
        filed_date = _date_or_none(row[H["Filed Date"]])

        # permit_id: prefer permit_number; otherwise derive county-date-idx
        if permit_number:
            permit_id = f"permit-{permit_number}"
        else:
            key = f"{county}-{filed_date or 'undated'}"
            derived_id_counter[key] = derived_id_counter.get(key, 0) + 1
            permit_id = f"permit-{county.lower().replace(' ', '-')}-{filed_date or 'undated'}-{derived_id_counter[key]}"

        shl_raw = row[H["Legal"]]
        bhl_raw = row[H["Bottom Hole Legal"]]
        shl_norm = parse_section_from_legal(shl_raw)
        bhl_norm = parse_section_from_legal(bhl_raw)
        sections = sorted({s for s in (shl_norm, bhl_norm) if s})

        matched_tract_ids: List[str] = []
        seen: set = set()
        for sec in sections:
            for tid in owned_index.get(sec, []):
                if tid not in seen:
                    matched_tract_ids.append(tid)
                    seen.add(tid)
        matched_tract_ids.sort()
        if matched_tract_ids:
            affecting_owned += 1

        purpose = to_str_or_none(row[H["Purpose Of Filing"]])
        doc_type = to_str_or_none(row[H["Type Of Document"]])
        permit_type = " — ".join([s for s in (doc_type, purpose) if s]) if (doc_type or purpose) else None

        permit_obj = {
            "approval_date": _date_or_none(row[H["Approval Date"]]),
            "approval_status": to_str_or_none(row[H["Approval Status"]]),
            "county": county,
            "effective_date": _date_or_none(row[H["Effective Date"]]),
            "lat": to_float_or_none(row[H["Latitude"]]),
            "lon": to_float_or_none(row[H["Longitude"]]),
            "matched_tract_ids": matched_tract_ids,
            "operator": to_str_or_none(row[H["Operator"]]),
            "oseberg_url": None,  # not in this Oseberg export
            "permit_date": filed_date,
            "permit_id": permit_id,
            "permit_number": permit_number,
            "permit_type": permit_type,
            "sections": sections,
            "source": to_str_or_none(row[H["Source"]]),
            "well_id": api_number,  # may be null per spec §6 amendment
            "well_name": to_str_or_none(row[H["Well Name"]]),
            "wellbore_profile": to_str_or_none(row[H["Wellbore Profile"]]),
        }
        permits.append(permit_obj)

    counters = {
        "permits_dropped_out_of_scope": dropped_out_of_scope,
        "permits_affecting_owned_tracts": affecting_owned,
    }
    return permits, counters, ingestion_errors


def merge_meta(meta_path: Path, *, source_folder: str, count: int, counters: dict, errors: List[dict]) -> dict:
    if meta_path.exists():
        meta = json.load(meta_path.open("r", encoding="utf-8"))
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}
    meta.setdefault("counts", {})["permits"] = count
    mr = meta.setdefault("matching_report", {})
    mr["permits_affecting_owned_tracts"] = counters["permits_affecting_owned_tracts"]
    mr["permits_with_owned_tract"] = counters["permits_affecting_owned_tracts"]  # legacy alias
    mr["permits_dropped_out_of_scope"] = counters["permits_dropped_out_of_scope"]
    existing_errors = mr.get("ingestion_errors", [])
    others = [e for e in existing_errors if e.get("section") != "permits"]
    mr["ingestion_errors"] = others + errors
    meta["generated_at"] = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not meta.get("oseberg_folder"):
        meta["oseberg_folder"] = source_folder
    return meta


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oseberg-folder", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    folder = OSEBERG_ROOT / args.oseberg_folder if args.oseberg_folder else find_latest_oseberg_folder(OSEBERG_ROOT)
    xlsx = folder / "permit_wab.xlsx"
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr)
        return 2

    owned_index = load_owned_index(TRACTS_IN)
    permits, counters, errors = parse_permits(xlsx, owned_index)
    permits.sort(key=lambda p: (p["permit_date"] or "0000-00-00", p["permit_id"]), reverse=True)

    now_utc = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    payload = {
        "generated_at": now_utc,
        "source": f"oseberg-{folder.name}",
        "permits": permits,
    }

    if not args.dry_run:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        json.dump(payload, PERMITS_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        PERMITS_OUT.open("a", encoding="utf-8").write("\n")
        meta = merge_meta(META_OUT, source_folder=folder.name, count=len(permits), counters=counters, errors=errors)
        json.dump(meta, META_OUT.open("w", encoding="utf-8"), indent=2, sort_keys=True, ensure_ascii=False)
        META_OUT.open("a", encoding="utf-8").write("\n")

    print(f"Permits in scope:                {len(permits)}")
    print(f"  affecting owned tracts:        {counters['permits_affecting_owned_tracts']}")
    print(f"Permits dropped (out of scope):  {counters['permits_dropped_out_of_scope']}")
    print(f"Ingestion errors:                {len(errors)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
