"""
Read data-raw/inventory/inventory-current.xlsx, normalize, assign tract IDs
via the registry, and write data/tracts.json and a partial data/meta.json.

Idempotent — running twice with no source changes produces identical output
and no registry updates.

See docs/2026-05-22-data-model-spec.md for the spec this script implements.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import shutil
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Make sibling import work when run as `python scripts/ingest_inventory.py`
SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import (  # noqa: E402
    NormalizationError,
    derive_orri_status_category,
    normalize_county,
    normalize_deal_slug,
    normalize_status_mineral,
    normalize_str,
    orri_row_hash,
    round_acres,
    to_iso_date_or_hbp,
    township_range_from_str,
)

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = SCRIPTS_DIR.parent
INVENTORY_PATH = PROJECT_ROOT / "data-raw" / "inventory" / "inventory-current.xlsx"
ARCHIVE_DIR = PROJECT_ROOT / "data-raw" / "inventory" / "archive"
REGISTRY_PATH = PROJECT_ROOT / "data-raw" / "inventory" / "id-registry.json"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_OUT = DATA_DIR / "tracts.json"
META_OUT = DATA_DIR / "meta.json"

SHEET_NAME = "Inventory 4 Sale"
# Documented ORRI columns (zero-indexed within the row tuple openpyxl yields):
# col index 2 = COUNTY, 3 = STR, 4 = NRA, 5 = DOL, 6 = EXP
# Anything in column index 8 (col I) is an aggregate-calculation cell per
# spec §13.1 and must be reported in meta.matching_report.aggregate_cells_skipped.
ORRI_AGGREGATE_COL_IDX = 8  # column I in 1-indexed terms

# Per-NRA valuation rates for ORRI tracts. These are Gib's documented pricing
# assumptions and match the aggregate cells I47 (non-HBP) and I92 (HBP) in
# the inventory workbook. Per spec §4.4, ORRIs of any other status_category
# get null asking fields. To change rates, edit here once.
ORRI_SALES_PER_NRA = {
    "HBP": 3500,
    "NON_PRODUCING": 1500,
}


# ---------------------------------------------------------------------------
# Inventory archive
# ---------------------------------------------------------------------------


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def archive_inventory(inventory_path: Path, archive_dir: Path) -> Path:
    """Snapshot the current inventory file into the archive.

    Naming: ``YYYY-MM-DD-inventory.xlsx`` based on the file's mtime date.
    If an archive with that name already exists:
      - identical content -> skip (return existing path).
      - different content -> append ``-2``, ``-3``, ... (per spec §13).
    """
    archive_dir.mkdir(parents=True, exist_ok=True)
    mtime_date = dt.date.fromtimestamp(inventory_path.stat().st_mtime).isoformat()
    base = f"{mtime_date}-inventory.xlsx"
    target = archive_dir / base
    if not target.exists():
        shutil.copy2(inventory_path, target)
        return target
    if _file_sha256(target) == _file_sha256(inventory_path):
        return target  # already archived, identical content
    # Same date, different content — append sequence suffix
    for n in range(2, 1000):
        candidate = archive_dir / f"{mtime_date}-inventory-{n}.xlsx"
        if not candidate.exists():
            shutil.copy2(inventory_path, candidate)
            return candidate
        if _file_sha256(candidate) == _file_sha256(inventory_path):
            return candidate
    raise RuntimeError("archive_inventory: ran out of sequence suffixes")


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------


def _is_blank(cells: Tuple[Any, ...]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in cells)


def _scan_for_header(ws, header_signature: Tuple[Tuple[int, str], ...]) -> int:
    """Find the row number whose cell at each (col_idx, expected_value) matches.

    col_idx is zero-indexed within the row tuple openpyxl yields. Returns the
    1-indexed worksheet row number, or raises if not found.
    """
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if all(
            col_idx < len(row)
            and isinstance(row[col_idx], str)
            and row[col_idx].strip().upper() == expected.upper()
            for col_idx, expected in header_signature
        ):
            return r_idx
    sig_str = ", ".join(f"col{i}={v!r}" for i, v in header_signature)
    raise RuntimeError(f"could not find header row matching: {sig_str}")


def _coerce_numeric(value, row_idx: int, col_label: str, section: str,
                    ingestion_errors: List[dict],
                    cast: str = "float") -> Optional[float]:
    """Convert a spreadsheet value to float/int or None, flagging non-numeric strings.

    'N/A', 'NA', and empty strings are silently treated as missing-data (null).
    Any other non-numeric value is recorded in ingestion_errors with the raw value
    preserved, and stored as null. Per spec §1.3 and CLAUDE.md Hard Rule #3.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if cast == "float" else int(round(value))
    if isinstance(value, str):
        s = value.strip()
        if not s or s.upper() in {"N/A", "NA", "-", "—"}:
            return None
        # Try to parse a numeric string ("30" -> 30, "0.1875" -> 0.1875)
        try:
            num = float(s)
            return num if cast == "float" else int(round(num))
        except ValueError:
            ingestion_errors.append(
                {
                    "section": section,
                    "row": row_idx,
                    "reason": f"non-numeric value in {col_label}: {value!r} — stored as null",
                    "raw": value,
                }
            )
            return None
    ingestion_errors.append(
        {
            "section": section,
            "row": row_idx,
            "reason": f"unexpected type for {col_label}: {type(value).__name__} — stored as null",
            "raw": str(value),
        }
    )
    return None


def _cell_hyperlink(cell) -> Optional[str]:
    """Return the URL embedded in a cell, or None if no hyperlink set."""
    if cell is None:
        return None
    link = getattr(cell, "hyperlink", None)
    if link is None:
        return None
    target = getattr(link, "target", None)
    return str(target) if target else None


def _row_is_all_blank(ws, r_idx: int, max_col: int) -> bool:
    """Tolerant blank-row check: True iff every cell in this row is None / empty string."""
    for c in range(1, max_col + 1):
        v = ws.cell(row=r_idx, column=c).value
        if v is not None and not (isinstance(v, str) and not v.strip()):
            return False
    return True


def parse_mineral_section(ws, ingestion_errors: List[dict]) -> List[dict]:
    """Find and parse the mineral section. Returns a list of raw-but-normalized row dicts.

    Uses header-name-driven column lookup so the parser is resilient to
    column insertions / reorderings in future inventory refreshes. The
    2026-05-22 schema (post inventory update) has columns:
      B=COUNTY  C=DEAL    D=STR    E=TR    F=NMA    G=ROYALTY
      H=STATUS  I=REGULATORY  J=Notes
      K=LEASE EXP  L=NRA  M=SALES PER NRA  N=SALES REVENUE  O=SALES PER NMA
    """
    header_row = _scan_for_header(ws, ((1, "COUNTY"), (2, "DEAL"), (3, "STR")))
    # Build header_name -> 1-indexed column number map (case-insensitive on the key)
    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            header_map[v.strip().upper()] = c

    def col(name: str, *, required: bool = True) -> Optional[int]:
        key = name.upper()
        if key not in header_map:
            if required:
                raise RuntimeError(f"mineral section: column {name!r} not found in header row {header_row}")
            return None
        return header_map[key]

    col_county = col("COUNTY")
    col_deal = col("DEAL")
    col_str = col("STR")
    col_nma = col("NMA")
    col_royalty = col("ROYALTY")
    col_status = col("STATUS")
    col_reg = col("REGULATORY", required=False)
    col_notes = col("NOTES", required=False)
    col_lease_exp = col("LEASE EXP")
    col_nra = col("NRA")
    col_sales_per_nra = col("SALES PER NRA")
    col_sales_revenue = col("SALES REVENUE")
    col_sales_per_nma = col("SALES PER NMA")

    rows: List[dict] = []
    r_idx = header_row + 1
    while r_idx <= ws.max_row:
        if _row_is_all_blank(ws, r_idx, ws.max_column):
            break  # end of mineral section
        county_cell = ws.cell(row=r_idx, column=col_county)
        if county_cell.value is None:
            r_idx += 1
            continue
        try:
            county = normalize_county(county_cell.value)
            str_canonical = normalize_str(ws.cell(row=r_idx, column=col_str).value)
            deal_raw = ws.cell(row=r_idx, column=col_deal).value
            deal_name = deal_raw if isinstance(deal_raw, str) else (str(deal_raw) if deal_raw is not None else "")
            deal_slug = normalize_deal_slug(deal_name)

            status_cell = ws.cell(row=r_idx, column=col_status)
            status_raw, status_category = normalize_status_mineral(status_cell.value)
            lease_url = _cell_hyperlink(status_cell)

            reg_cell = ws.cell(row=r_idx, column=col_reg) if col_reg else None
            regulatory_status = reg_cell.value if reg_cell and reg_cell.value else None
            regulatory_url = _cell_hyperlink(reg_cell) if reg_cell else None

            notes_cell = ws.cell(row=r_idx, column=col_notes) if col_notes else None
            notes = notes_cell.value if notes_cell and notes_cell.value else None
            if isinstance(notes, str):
                notes = notes.strip() or None

            lease_exp = to_iso_date_or_hbp(ws.cell(row=r_idx, column=col_lease_exp).value)
            nma = _coerce_numeric(ws.cell(row=r_idx, column=col_nma).value, r_idx, "NMA", "mineral", ingestion_errors)
            royalty = _coerce_numeric(ws.cell(row=r_idx, column=col_royalty).value, r_idx, "ROYALTY", "mineral", ingestion_errors)
            nra = _coerce_numeric(ws.cell(row=r_idx, column=col_nra).value, r_idx, "NRA", "mineral", ingestion_errors)
            sales_per_nra = _coerce_numeric(ws.cell(row=r_idx, column=col_sales_per_nra).value, r_idx, "SALES PER NRA", "mineral", ingestion_errors, cast="int")
            sales_revenue = _coerce_numeric(ws.cell(row=r_idx, column=col_sales_revenue).value, r_idx, "SALES REVENUE", "mineral", ingestion_errors, cast="int")
            sales_per_nma = _coerce_numeric(ws.cell(row=r_idx, column=col_sales_per_nma).value, r_idx, "SALES PER NMA", "mineral", ingestion_errors, cast="int")
            rows.append(
                {
                    "row_number": r_idx,
                    "type": "mineral",
                    "county": county,
                    "str": str_canonical,
                    "township_range": township_range_from_str(str_canonical),
                    "deal_name": deal_name,
                    "deal_slug": deal_slug,
                    "nma": round_acres(nma),
                    "royalty": royalty,
                    "status_raw": status_raw,
                    "status_category": status_category,
                    "lease_url": lease_url,
                    "regulatory_status": regulatory_status,
                    "regulatory_url": regulatory_url,
                    "notes": notes,
                    "lease_expiration": lease_exp,
                    "nra": round_acres(nra),
                    "sales_per_nra": sales_per_nra,
                    "sales_revenue": sales_revenue,
                    "sales_per_nma": sales_per_nma,
                }
            )
        except NormalizationError as exc:
            ingestion_errors.append(
                {"section": "mineral", "row": r_idx, "reason": str(exc)}
            )
        r_idx += 1
    return rows


def parse_orri_section(
    ws, ingestion_errors: List[dict], aggregate_cells_skipped: List[dict]
) -> List[dict]:
    """Find and parse the ORRI section. Returns a list of normalized row dicts.

    Header-driven column lookup. 2026-05-22 ORRI schema:
      C=COUNTY  D=STR  E=NRA  F=DOL  G=EXP  I=REGULATORY  J=Notes
    Aggregate-valuation cells (per-acre rates, projected sale-price totals)
    now sit in column K (was column I before the schema upgrade).
    """
    header_row = _scan_for_header(ws, ((2, "COUNTY"), (3, "STR"), (4, "NRA")))
    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            header_map[v.strip().upper()] = c

    col_county = header_map["COUNTY"]
    col_str = header_map["STR"]
    col_nra = header_map["NRA"]
    col_dol = header_map["DOL"]
    col_exp = header_map["EXP"]
    col_reg = header_map.get("REGULATORY")
    col_notes = header_map.get("NOTES")

    # Aggregate cells now in column K (1-indexed col 11). Any non-null cell
    # in this column inside the ORRI section is recorded as an aggregate cell
    # in the matching report. (Per spec §13.1.)
    AGGREGATE_COL = 11
    documented_orri_cols = {
        col_county, col_str, col_nra, col_dol, col_exp,
    }
    if col_reg: documented_orri_cols.add(col_reg)
    if col_notes: documented_orri_cols.add(col_notes)

    rows: List[dict] = []
    r_idx = header_row + 1
    # Occurrence counter per identity tuple (county, str, nra, dol, exp).
    # Spec §3.4: 1st occurrence uses 5-field hash (counter=0, no suffix in
    # hash key); 2nd+ occurrences append the counter to differentiate.
    tuple_counter: Dict[Tuple[str, str, Optional[float], Optional[str], Optional[str]], int] = {}

    while r_idx <= ws.max_row:
        county_cell = ws.cell(row=r_idx, column=col_county)
        is_data_row = county_cell.value is not None

        # Capture any aggregate-cell value on this row (with or without data)
        agg_cell = ws.cell(row=r_idx, column=AGGREGATE_COL)
        if agg_cell.value is not None:
            aggregate_cells_skipped.append(
                {
                    "sheet": SHEET_NAME,
                    "cell": f"{get_column_letter(AGGREGATE_COL)}{r_idx}",
                    "value": agg_cell.value,
                    "interpretation": "ORRI section column-K aggregate calculation",
                }
            )

        if not is_data_row:
            r_idx += 1
            continue

        try:
            county = normalize_county(county_cell.value)
            str_canonical = normalize_str(ws.cell(row=r_idx, column=col_str).value)
            nra_raw = _coerce_numeric(ws.cell(row=r_idx, column=col_nra).value, r_idx, "NRA", "orri", ingestion_errors)
            nra = round_acres(nra_raw)
            dol = to_iso_date_or_hbp(ws.cell(row=r_idx, column=col_dol).value)
            exp = to_iso_date_or_hbp(ws.cell(row=r_idx, column=col_exp).value)
            status_category = derive_orri_status_category(dol, exp)
            identity_tuple = (county, str_canonical, nra, dol, exp)
            occurrence = tuple_counter.get(identity_tuple, 0)
            tuple_counter[identity_tuple] = occurrence + 1
            row_hash = orri_row_hash(county, str_canonical, nra, dol, exp, occurrence)

            reg_cell = ws.cell(row=r_idx, column=col_reg) if col_reg else None
            regulatory_status = reg_cell.value if reg_cell and reg_cell.value else None
            regulatory_url = _cell_hyperlink(reg_cell) if reg_cell else None
            notes_cell = ws.cell(row=r_idx, column=col_notes) if col_notes else None
            notes = notes_cell.value if notes_cell and notes_cell.value else None
            if isinstance(notes, str):
                notes = notes.strip() or None

            rows.append(
                {
                    "row_number": r_idx,
                    "type": "orri",
                    "county": county,
                    "str": str_canonical,
                    "township_range": township_range_from_str(str_canonical),
                    "nra": nra,
                    "status_raw": None,
                    "status_category": status_category,
                    "date_of_lease": dol,
                    "lease_expiration": exp,
                    "row_hash": row_hash,
                    "regulatory_status": regulatory_status,
                    "regulatory_url": regulatory_url,
                    "notes": notes,
                }
            )
            if status_category == "OTHER":
                ingestion_errors.append(
                    {
                        "section": "orri",
                        "row": r_idx,
                        "reason": f"unexpected DOL/EXP combination: DOL={dol!r}, EXP={exp!r}",
                    }
                )
        except NormalizationError as exc:
            ingestion_errors.append(
                {"section": "orri", "row": r_idx, "reason": str(exc)}
            )
        r_idx += 1

    # URL propagation: for ORRI rows where multiple share the same
    # (str, regulatory_status) but only one has a regulatory_url, propagate
    # that URL to siblings. Same regulatory case applies to all of them.
    url_map: Dict[Tuple[str, str], str] = {}
    for r in rows:
        rs = r.get("regulatory_status")
        ru = r.get("regulatory_url")
        if rs and ru:
            url_map[(r["str"], rs)] = ru
    propagated = 0
    for r in rows:
        rs = r.get("regulatory_status")
        if rs and not r.get("regulatory_url"):
            key = (r["str"], rs)
            if key in url_map:
                r["regulatory_url"] = url_map[key]
                propagated += 1
    if propagated:
        # Informational only — not an ingestion error. Print so it shows in the
        # console output during refresh runs.
        print(f"  ORRI: propagated regulatory_url to {propagated} sibling rows "
              f"(same STR + regulatory_status, original link on one row only)")

    return rows


# ---------------------------------------------------------------------------
# ID registry
# ---------------------------------------------------------------------------


def load_registry(path: Path) -> dict:
    if not path.exists():
        return {"minerals": {}, "orri": {}, "retired": []}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_registry(path: Path, registry: dict) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, sort_keys=True, ensure_ascii=False)
        f.write("\n")


def _next_id(prefix: str, existing_ids: set) -> str:
    """Return the next sequential ID (e.g., Min001) not in existing_ids."""
    nums = []
    for k in existing_ids:
        if k.startswith(prefix):
            try:
                nums.append(int(k[len(prefix):]))
            except ValueError:
                continue
    nxt = (max(nums) + 1) if nums else 1
    return f"{prefix}{nxt:03d}"


def assign_mineral_ids(rows: List[dict], registry: dict, today_iso: str) -> Tuple[List[dict], List[str], List[str]]:
    """Assign Min IDs to mineral rows. Returns (rows_with_ids, newly_assigned, newly_retired)."""
    # Sort canonically per spec §3.2
    rows_sorted = sorted(rows, key=lambda r: (r["county"], r["str"], r["deal_slug"]))

    # Build lookup: (county, str, deal_slug) -> mineral_id from registry
    minerals = registry.setdefault("minerals", {})
    retired = registry.setdefault("retired", [])
    reverse_lookup = {
        (entry["county"], entry["str"], entry["deal_slug"]): tid
        for tid, entry in minerals.items()
    }

    seen_identities: set = set()
    newly_assigned: List[str] = []

    for row in rows_sorted:
        identity = (row["county"], row["str"], row["deal_slug"])
        seen_identities.add(identity)
        if identity in reverse_lookup:
            row["tract_id"] = reverse_lookup[identity]
            row["first_seen"] = minerals[row["tract_id"]].get("first_seen", today_iso)
        else:
            all_used = set(minerals.keys()) | {e["tract_id"] for e in retired if e.get("tract_id", "").startswith("Min")}
            new_id = _next_id("Min", all_used)
            row["tract_id"] = new_id
            row["first_seen"] = today_iso
            minerals[new_id] = {
                "county": row["county"],
                "str": row["str"],
                "deal_slug": row["deal_slug"],
                "first_seen": today_iso,
            }
            newly_assigned.append(new_id)

    # Detect retired (in registry but not in current inventory)
    newly_retired: List[str] = []
    for tid, entry in list(minerals.items()):
        identity = (entry["county"], entry["str"], entry["deal_slug"])
        if identity not in seen_identities:
            retired.append({"tract_id": tid, **entry, "retired_on": today_iso})
            del minerals[tid]
            newly_retired.append(tid)

    return rows_sorted, newly_assigned, newly_retired


def assign_orri_ids(rows: List[dict], registry: dict, today_iso: str) -> Tuple[List[dict], List[str], List[str]]:
    """Assign OR IDs to ORRI rows. Returns (rows_with_ids, newly_assigned, newly_retired)."""
    rows_sorted = sorted(rows, key=lambda r: (r["county"], r["str"], r["row_hash"]))

    orri = registry.setdefault("orri", {})
    retired = registry.setdefault("retired", [])
    reverse_lookup = {
        (entry["county"], entry["str"], entry["row_hash"]): tid
        for tid, entry in orri.items()
    }

    seen_identities: set = set()
    newly_assigned: List[str] = []

    for row in rows_sorted:
        identity = (row["county"], row["str"], row["row_hash"])
        seen_identities.add(identity)
        if identity in reverse_lookup:
            row["tract_id"] = reverse_lookup[identity]
            row["first_seen"] = orri[row["tract_id"]].get("first_seen", today_iso)
        else:
            all_used = set(orri.keys()) | {e["tract_id"] for e in retired if e.get("tract_id", "").startswith("OR")}
            new_id = _next_id("OR", all_used)
            row["tract_id"] = new_id
            row["first_seen"] = today_iso
            orri[new_id] = {
                "county": row["county"],
                "str": row["str"],
                "row_hash": row["row_hash"],
                "first_seen": today_iso,
            }
            newly_assigned.append(new_id)

    newly_retired: List[str] = []
    for tid, entry in list(orri.items()):
        identity = (entry["county"], entry["str"], entry["row_hash"])
        if identity not in seen_identities:
            retired.append({"tract_id": tid, **entry, "retired_on": today_iso})
            del orri[tid]
            newly_retired.append(tid)

    return rows_sorted, newly_assigned, newly_retired


# ---------------------------------------------------------------------------
# Tract object construction
# ---------------------------------------------------------------------------


def build_mineral_tract_obj(row: dict) -> dict:
    """Per spec §4.2 + §4.3. Keys sorted for git-diff stability (§1.4)."""
    return {
        "county": row["county"],
        "deal_name": row["deal_name"],
        "deal_slug": row["deal_slug"],
        "first_seen": row["first_seen"],
        "lat": None,
        "lease_expiration": row["lease_expiration"],
        "lease_url": row.get("lease_url"),
        "lon": None,
        "nma": row["nma"],
        "notes": row.get("notes"),
        "nra": row["nra"],
        "regulatory_status": row.get("regulatory_status"),
        "regulatory_url": row.get("regulatory_url"),
        "royalty": row["royalty"],
        "sales_per_nma": row["sales_per_nma"],
        "sales_per_nra": row["sales_per_nra"],
        "sales_revenue": row["sales_revenue"],
        "section_polygon": None,
        "status_category": row["status_category"],
        "status_raw": row["status_raw"],
        "str": row["str"],
        "township_range": row["township_range"],
        "tract_id": row["tract_id"],
        "type": "mineral",
    }


def build_orri_tract_obj(row: dict) -> dict:
    """Per spec §4.2 + §4.4. Keys sorted for git-diff stability (§1.4)."""
    sales_per_nra = ORRI_SALES_PER_NRA.get(row["status_category"])
    if sales_per_nra is not None and row.get("nra") is not None:
        sales_revenue = int(round(row["nra"] * sales_per_nra))
    else:
        sales_revenue = None
    return {
        "county": row["county"],
        "date_of_lease": row["date_of_lease"],
        "first_seen": row["first_seen"],
        "lat": None,
        "lease_expiration": row["lease_expiration"],
        "lon": None,
        "notes": row.get("notes"),
        "nra": row["nra"],
        "regulatory_status": row.get("regulatory_status"),
        "regulatory_url": row.get("regulatory_url"),
        "row_hash": row["row_hash"],
        "sales_per_nra": sales_per_nra,
        "sales_revenue": sales_revenue,
        "section_polygon": None,
        "status_category": row["status_category"],
        "status_raw": None,
        "str": row["str"],
        "township_range": row["township_range"],
        "tract_id": row["tract_id"],
        "type": "orri",
    }


def _tract_sort_key(t: dict) -> Tuple[str, int]:
    """Sort by type then numeric ID. Min before OR. Within each, numeric order."""
    type_rank = 0 if t["type"] == "mineral" else 1
    prefix = "Min" if t["type"] == "mineral" else "OR"
    try:
        n = int(t["tract_id"][len(prefix):])
    except ValueError:
        n = 0
    return (type_rank, n)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report counts; do not write tracts.json / meta.json / registry.",
    )
    args = parser.parse_args(argv)

    if not INVENTORY_PATH.exists():
        print(f"ERROR: inventory file not found at {INVENTORY_PATH}", file=sys.stderr)
        return 2

    today_iso = dt.date.today().isoformat()
    now_utc = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    # 1. Archive the current inventory file (snapshot)
    archive_path = archive_inventory(INVENTORY_PATH, ARCHIVE_DIR)

    # 2. Parse the workbook
    wb = load_workbook(INVENTORY_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet {SHEET_NAME!r} not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 3
    ws = wb[SHEET_NAME]

    ingestion_errors: List[dict] = []
    aggregate_cells_skipped: List[dict] = []

    mineral_rows = parse_mineral_section(ws, ingestion_errors)
    orri_rows = parse_orri_section(ws, ingestion_errors, aggregate_cells_skipped)

    # 3. ID assignment
    registry = load_registry(REGISTRY_PATH)
    mineral_rows, new_min, ret_min = assign_mineral_ids(mineral_rows, registry, today_iso)
    orri_rows, new_orri, ret_orri = assign_orri_ids(orri_rows, registry, today_iso)

    # 4. Build tract objects, sort, and assemble tracts.json payload
    tracts = [build_mineral_tract_obj(r) for r in mineral_rows] + [build_orri_tract_obj(r) for r in orri_rows]
    tracts.sort(key=_tract_sort_key)

    tracts_payload = {
        "generated_at": now_utc,
        "source_inventory": INVENTORY_PATH.name,
        "source_inventory_archived_as": archive_path.name,
        "tracts": tracts,
    }

    # 5. Build partial meta.json payload (counts + id_registry + matching_report shell)
    meta_payload = {
        "generated_at": now_utc,
        "inventory_file": INVENTORY_PATH.name,
        "inventory_archived_as": archive_path.name,
        "oseberg_folder": None,
        "counts": {
            "tracts_mineral": len(mineral_rows),
            "tracts_orri": len(orri_rows),
            "wells": 0,
            "permits": 0,
            "completions": 0,
            "production_records": 0,
            "leases": 0,
            "regulatory_actions": 0,
        },
        "matching_report": {
            "wells_with_owned_tract": 0,
            "wells_without_owned_tract": 0,
            "permits_with_owned_tract": 0,
            "leases_affecting_owned_tracts": 0,
            "regulatory_affecting_owned_tracts": 0,
            "ingestion_errors": ingestion_errors,
            "aggregate_cells_skipped": aggregate_cells_skipped,
        },
        "id_registry": {
            "minerals_assigned": len(registry.get("minerals", {})),
            "minerals_retired": sum(1 for e in registry.get("retired", []) if e.get("tract_id", "").startswith("Min")),
            "orri_assigned": len(registry.get("orri", {})),
            "orri_retired": sum(1 for e in registry.get("retired", []) if e.get("tract_id", "").startswith("OR")),
            "new_this_refresh": new_min + new_orri,
            "retired_this_refresh": ret_min + ret_orri,
        },
    }

    # 6. Write outputs
    if args.dry_run:
        print("DRY RUN — no files written.")
    else:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with TRACTS_OUT.open("w", encoding="utf-8") as f:
            json.dump(tracts_payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")
        with META_OUT.open("w", encoding="utf-8") as f:
            json.dump(meta_payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")
        save_registry(REGISTRY_PATH, registry)

    # 7. Console summary
    print("=" * 60)
    print(f"Inventory:       {INVENTORY_PATH.name}")
    print(f"Archive path:    {archive_path.relative_to(PROJECT_ROOT)}")
    print(f"Mineral tracts:  {len(mineral_rows)}")
    print(f"ORRI tracts:     {len(orri_rows)}")
    print(f"  HBP ORRIs:           {sum(1 for r in orri_rows if r['status_category'] == 'HBP')}")
    print(f"  NON_PRODUCING ORRIs: {sum(1 for r in orri_rows if r['status_category'] == 'NON_PRODUCING')}")
    print(f"  OTHER ORRIs:         {sum(1 for r in orri_rows if r['status_category'] == 'OTHER')}")
    print(f"Newly assigned:  {len(new_min) + len(new_orri)}  ({', '.join(new_min + new_orri) or 'none'})")
    print(f"Newly retired:   {len(ret_min) + len(ret_orri)}  ({', '.join(ret_min + ret_orri) or 'none'})")
    print(f"Ingestion errs:  {len(ingestion_errors)}")
    print(f"Aggregate cells skipped: {len(aggregate_cells_skipped)}  ({', '.join(c['cell'] for c in aggregate_cells_skipped) or 'none'})")
    print("=" * 60)

    return 1 if ingestion_errors else 0


if __name__ == "__main__":
    sys.exit(main())
