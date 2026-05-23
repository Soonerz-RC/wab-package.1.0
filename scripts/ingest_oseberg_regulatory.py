"""
Read three Oseberg sources and unify them into data/regulatory.json:
  - pooling_wab.xlsx  -> POOLING_APPLICATION | POOLING_ORDER
  - spacing_wab.xlsx  -> SPACING_APPLICATION | SPACING_ORDER
  - le_wab.xlsx       -> LOCATION_EXCEPTION

See docs/2026-05-22-data-model-spec.md §10.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from normalize import (  # noqa: E402
    CANONICAL_COUNTIES,
    NormalizationError,
    normalize_county,
    normalize_upland_in_string,
)
from ingest_oseberg_wells import (  # noqa: E402
    find_latest_oseberg_folder,
    parse_section_from_legal,
    to_str_or_none,
    _date_or_none,
)
from openpyxl import load_workbook  # noqa: E402

PROJECT_ROOT = SCRIPTS_DIR.parent
OSEBERG_ROOT = PROJECT_ROOT / "data-raw" / "oseberg"
DATA_DIR = PROJECT_ROOT / "data"
TRACTS_IN = DATA_DIR / "tracts.json"
REGULATORY_OUT = DATA_DIR / "regulatory.json"
META_OUT = DATA_DIR / "meta.json"


def load_owned_index(tracts_path: Path) -> Dict[str, List[str]]:
    payload = json.load(tracts_path.open("r", encoding="utf-8"))
    idx: Dict[str, List[str]] = {}
    for t in payload.get("tracts", []):
        idx.setdefault(t["str"], []).append(t["tract_id"])
    for k in idx:
        idx[k].sort()
    return idx


def _classify_pooling(app_or_order: Optional[str]) -> str:
    if app_or_order and app_or_order.strip().lower().startswith("app"):
        return "POOLING_APPLICATION"
    return "POOLING_ORDER"


def _classify_spacing(app_or_order: Optional[str]) -> str:
    if app_or_order and app_or_order.strip().lower().startswith("app"):
        return "SPACING_APPLICATION"
    return "SPACING_ORDER"


def _parse_common(
    row, H: Dict[str, int], owned_index: Dict[str, List[str]], r_idx: int,
) -> Tuple[Optional[str], List[str], List[str], bool, str]:
    """Return (county_or_None, sections, matched_tract_ids, affects_owned, county_canonical_or_empty)."""
    county_raw = row[H["County"]]
    if county_raw is None:
        return None, [], [], False, ""
    try:
        county = normalize_county(county_raw)
    except NormalizationError:
        return None, [], [], False, ""
    if county not in CANONICAL_COUNTIES:
        return None, [], [], False, ""

    legal_raw = to_str_or_none(row[H["Legal"]]) if "Legal" in H else None
    section_norm = parse_section_from_legal(legal_raw)
    sections = [section_norm] if section_norm else []

    matched: List[str] = []
    if sections:
        for tid in owned_index.get(sections[0], []):
            matched.append(tid)
    matched.sort()
    return county, sections, matched, bool(matched), county


def _parse_pooling(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], int, int]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    records: List[dict] = []
    dropped = 0
    affecting_owned = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county, sections, matched, affects, _ = _parse_common(row, H, owned_index, r_idx)
        if county is None:
            dropped += 1
            continue

        app_or_order = to_str_or_none(row[H["App or Order"]]) if "App or Order" in H else None
        cause_number = to_str_or_none(row[H["Cause Number"]]) if "Cause Number" in H else None
        action_type = _classify_pooling(app_or_order)
        action_id = f"pool-{cause_number or 'unknown'}-{action_type.lower()}-r{r_idx}"

        applicant = normalize_upland_in_string(to_str_or_none(row[H["Applicant"]]) if "Applicant" in H else None)
        formation = to_str_or_none(row[H["Formation"]]) if "Formation" in H else None
        legal_raw = to_str_or_none(row[H["Legal"]]) if "Legal" in H else None

        summary_parts = []
        summary_parts.append(f"{action_type.replace('_', ' ').title()}")
        if legal_raw:
            summary_parts.append(f"on {legal_raw}")
        if formation:
            summary_parts.append(f"({formation})")
        if applicant:
            summary_parts.append(f"by {applicant}")
        summary = " ".join(summary_parts)

        record = {
            "action_id": action_id,
            "affects_owned_tracts": affects,
            "applicant": applicant,
            "cause_number": cause_number,
            "county": county,
            "effective_date": _date_or_none(row[H["Order Date"]]) if "Order Date" in H else None,
            "filing_date": _date_or_none(row[H["App Date"]]) if "App Date" in H else None,
            "matched_tract_ids": matched,
            "oseberg_url": None,  # spec §10 amended: not in exports
            "oseberg_url_requires_login": True,
            "raw": {
                "app_or_order": app_or_order,
                "formation": formation,
                "order_number": to_str_or_none(row[H["Order Number"]]) if "Order Number" in H else None,
                "order_type": to_str_or_none(row[H["Order Type"]]) if "Order Type" in H else None,
                "configuration": to_str_or_none(row[H["Configuration"]]) if "Configuration" in H else None,
                "unit_size": to_str_or_none(row[H["Unit Size"]]) if "Unit Size" in H else None,
                "status": to_str_or_none(row[H["Status"]]) if "Status" in H else None,
                "horizontal": to_str_or_none(row[H["Horizontal"]]) if "Horizontal" in H else None,
            },
            "sections": sections,
            "source_file": "pooling_wab.xlsx",
            "summary": summary,
            "type": action_type,
        }
        records.append(record)
        if affects:
            affecting_owned += 1

    return records, dropped, affecting_owned


def _parse_spacing(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], int, int]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    records: List[dict] = []
    dropped = 0
    affecting_owned = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county, sections, matched, affects, _ = _parse_common(row, H, owned_index, r_idx)
        if county is None:
            dropped += 1
            continue

        app_or_order = to_str_or_none(row[H["App or Order"]]) if "App or Order" in H else None
        cause_number = to_str_or_none(row[H["Cause Number"]]) if "Cause Number" in H else None
        action_type = _classify_spacing(app_or_order)
        action_id = f"space-{cause_number or 'unknown'}-{action_type.lower()}-r{r_idx}"

        applicant = normalize_upland_in_string(to_str_or_none(row[H["Applicant"]]) if "Applicant" in H else None)
        formation = to_str_or_none(row[H["Formation"]]) if "Formation" in H else None
        legal_raw = to_str_or_none(row[H["Legal"]]) if "Legal" in H else None

        summary_parts = [f"{action_type.replace('_', ' ').title()}"]
        if legal_raw:
            summary_parts.append(f"on {legal_raw}")
        if formation:
            summary_parts.append(f"({formation})")
        if applicant:
            summary_parts.append(f"by {applicant}")
        summary = " ".join(summary_parts)

        record = {
            "action_id": action_id,
            "affects_owned_tracts": affects,
            "applicant": applicant,
            "cause_number": cause_number,
            "county": county,
            "effective_date": _date_or_none(row[H["Order Date"]]) if "Order Date" in H else None,
            "filing_date": _date_or_none(row[H["App Date"]]) if "App Date" in H else None,
            "matched_tract_ids": matched,
            "oseberg_url": None,
            "oseberg_url_requires_login": True,
            "raw": {
                "app_or_order": app_or_order,
                "formation": formation,
                "order_number": to_str_or_none(row[H["Order Number"]]) if "Order Number" in H else None,
                "order_type": to_str_or_none(row[H["Order Type"]]) if "Order Type" in H else None,
                "configuration": to_str_or_none(row[H["Configuration"]]) if "Configuration" in H else None,
                "unit_size": to_str_or_none(row[H["Unit Size"]]) if "Unit Size" in H else None,
                "status": to_str_or_none(row[H["Status"]]) if "Status" in H else None,
                "horizontal": to_str_or_none(row[H["Horizontal"]]) if "Horizontal" in H else None,
                "document_type": to_str_or_none(row[H["Document Type"]]) if "Document Type" in H else None,
            },
            "sections": sections,
            "source_file": "spacing_wab.xlsx",
            "summary": summary,
            "type": action_type,
        }
        records.append(record)
        if affects:
            affecting_owned += 1

    return records, dropped, affecting_owned


def _parse_le(xlsx_path: Path, owned_index: Dict[str, List[str]]) -> Tuple[List[dict], int, int]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(headers) if h is not None}

    records: List[dict] = []
    dropped = 0
    affecting_owned = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        county, sections, matched, affects, _ = _parse_common(row, H, owned_index, r_idx)
        if county is None:
            dropped += 1
            continue

        app_or_order = to_str_or_none(row[H["App or Order"]]) if "App or Order" in H else None
        cause_number = to_str_or_none(row[H["Cause Number"]]) if "Cause Number" in H else None
        action_id = f"le-{cause_number or 'unknown'}-r{r_idx}"

        applicant = normalize_upland_in_string(to_str_or_none(row[H["Applicant"]]) if "Applicant" in H else None)
        formation = to_str_or_none(row[H["Formation"]]) if "Formation" in H else None
        legal_raw = to_str_or_none(row[H["Legal"]]) if "Legal" in H else None

        summary_parts = ["Location Exception"]
        if legal_raw:
            summary_parts.append(f"on {legal_raw}")
        if formation:
            summary_parts.append(f"({formation})")
        if applicant:
            summary_parts.append(f"by {applicant}")
        summary = " ".join(summary_parts)

        record = {
            "action_id": action_id,
            "affects_owned_tracts": affects,
            "applicant": applicant,
            "cause_number": cause_number,
            "county": county,
            "effective_date": _date_or_none(row[H["Order Date"]]) if "Order Date" in H else None,
            "filing_date": _date_or_none(row[H["App Date"]]) if "App Date" in H else None,
            "matched_tract_ids": matched,
            "oseberg_url": None,
            "oseberg_url_requires_login": True,
            "raw": {
                "app_or_order": app_or_order,
                "formation": formation,
                "order_number": to_str_or_none(row[H["Order Number"]]) if "Order Number" in H else None,
                "order_type": to_str_or_none(row[H["Order Type"]]) if "Order Type" in H else None,
                "configuration": to_str_or_none(row[H["Configuration"]]) if "Configuration" in H else None,
                "unit_size": to_str_or_none(row[H["Unit Size"]]) if "Unit Size" in H else None,
                "status": to_str_or_none(row[H["Status"]]) if "Status" in H else None,
                "horizontal": to_str_or_none(row[H["Horizontal"]]) if "Horizontal" in H else None,
                "spacing_exception_order_number": to_str_or_none(row[H["Spacing Exception Order Number"]]) if "Spacing Exception Order Number" in H else None,
            },
            "sections": sections,
            "source_file": "le_wab.xlsx",
            "summary": summary,
            "type": "LOCATION_EXCEPTION",
        }
        records.append(record)
        if affects:
            affecting_owned += 1

    return records, dropped, affecting_owned


def merge_meta(meta_path: Path, *, source_folder: str, count: int, affecting_owned: int, dropped: int) -> dict:
    if meta_path.exists():
        meta = json.load(meta_path.open("r", encoding="utf-8"))
    else:
        meta = {"counts": {}, "matching_report": {}, "id_registry": {}}
    meta.setdefault("counts", {})["regulatory_actions"] = count
    mr = meta.setdefault("matching_report", {})
    mr["regulatory_affecting_owned_tracts"] = affecting_owned
    mr["regulatory_dropped_out_of_scope"] = dropped
    meta["generated_at"] = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    if not meta.get("oseberg_folder"):
        meta["oseberg_folder"] = source_folder
    return meta


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--oseberg-folder", type=str, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    folder = OSEBERG_ROOT / args.oseberg_folder if args.oseberg_folder else find_latest_oseberg_folder(OSEBERG_ROOT)
    pooling_xlsx = folder / "pooling_wab.xlsx"
    spacing_xlsx = folder / "spacing_wab.xlsx"
    le_xlsx = folder / "le_wab.xlsx"

    for p in (pooling_xlsx, spacing_xlsx, le_xlsx):
        if not p.exists():
            print(f"ERROR: {p} not found", file=sys.stderr)
            return 2

    owned_index = load_owned_index(TRACTS_IN)
    print(f"Parsing pooling_wab.xlsx...", flush=True)
    pooling, p_dropped, p_owned = _parse_pooling(pooling_xlsx, owned_index)
    print(f"Parsing spacing_wab.xlsx...", flush=True)
    spacing, s_dropped, s_owned = _parse_spacing(spacing_xlsx, owned_index)
    print(f"Parsing le_wab.xlsx...", flush=True)
    le, l_dropped, l_owned = _parse_le(le_xlsx, owned_index)

    all_records = pooling + spacing + le
    all_records.sort(
        key=lambda r: (r["filing_date"] or "0000-00-00", r["action_id"]),
        reverse=True,
    )

    payload = {
        "generated_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "source": f"oseberg-{folder.name}",
        "actions": all_records,
    }

    total_count = len(all_records)
    total_owned = p_owned + s_owned + l_owned
    total_dropped = p_dropped + s_dropped + l_dropped

    if not args.dry_run:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with REGULATORY_OUT.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")
        meta = merge_meta(META_OUT, source_folder=folder.name, count=total_count, affecting_owned=total_owned, dropped=total_dropped)
        with META_OUT.open("w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, sort_keys=True, ensure_ascii=False)
            f.write("\n")

    print()
    print(f"  Pooling:    {len(pooling):>5d} records  ({p_owned} affecting owned)")
    print(f"  Spacing:    {len(spacing):>5d} records  ({s_owned} affecting owned)")
    print(f"  Location Exception: {len(le):>5d} records  ({l_owned} affecting owned)")
    print(f"  ----")
    print(f"  Total:      {total_count:>5d} records  ({total_owned} affecting owned)")
    print(f"Dropped out-of-scope: {total_dropped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
